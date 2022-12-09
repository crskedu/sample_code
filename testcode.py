import json
from collections import OrderedDict
import os
#import system
import json
import numpy as np
from datetime import date
from datetime import datetime
import time

import smtplib, ssl
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText

import socket
import requests
import re
from bs4 import BeautifulSoup
import json
import collections
import sys
import csv
import openpyxl
import cx_Oracle
#src_ggadmin_user_name_v = ''

'''    
# test this    pavan.vachhani@oracle.com
def multiprocess():

    from multiprocessing import Process
    from time import sleep
    import time

def f(name):
    print('hello', name)
    time.sleep(10)
if __name__ == "__main__":
  start=time.perf_counter()
  a=Process(target=f, args=('hello',))
  b=Process(target=f, args=('stock',))
  a.start()
  b.start()
  a.join()
  b.join()
  finish=time.perf_counter()

  print("Finished in", {round(finish-start, 2)}, "seconds")

print(type(f))
print(type(f('ABC')))    
'''
    
def spool2text():
    left_alignment = "Left Text"
    center_alignment = "Centered"
    right_alignment = "Right Text"

    sys.stdout = open(r'C:\MyFolder\SE_py\file.txt', 'w')
    print('test')
    print('test2')
    print(f"{left_alignment : <10}|{center_alignment : ^15}|{right_alignment : >20}|")
    print(f"{center_alignment : ^15}|")
    print(f"{right_alignment : >20}|")    
    sys.stdout.close()

#https://www.askpython.com/python/built-in-methods/python-print-to-file
#https://www.askpython.com/python-modules/python-logging-module

    with open(r"C:\MyFolder\SE_py\output.txt", "a") as f:
        print('Hi', file=f)
        print('Hello from AskPython', file=f)
        print('exit', file=f)
        print('test',file=f)
        print('test2',file=f)
        print(f"{left_alignment : <10}|{center_alignment : ^15}|{right_alignment : >20}|",file=f)
        print(f"{center_alignment : ^15}|",file=f)
        print(f"{right_alignment : >20}|",file=f)    
        
    f.close()    
    
    
def printalign():
# https://medium.com/@NirantK/best-of-python3-6-f-strings-41f9154983e
# https://stackoverflow.com/questions/33594958/is-it-possible-to-align-a-print-statement-to-the-center-in-python
# assigning strings to the variables
    left_alignment = "Left Text"
    center_alignment = "Centered"
    right_alignment = "Right Text"

    # printing out aligned text
    print(f"{left_alignment : <10}|{center_alignment : ^15}|{right_alignment : >20}|")
    print(f"{center_alignment : ^15}|")
    print(f"{right_alignment : >20}|")
    print("Centre is %s right is %s" % center_alignment,'stt')
    print("Hello %s, my name is %s" % ('john', 'mike')) # %s is to convert string of any valye 
    # Hello john, my name is mike".
    print("My name is %s and i'm %d" % ('john', 12))    
    #My name is john and i'm 12          

def splitword():
    '''
    vlistname=['oracleidentitycloudservice/sathishkumar.rangaraj@oracle.com',
                'oracleidentitycloudservice/roopesh.thokala@oracle.com',
                'oracleidentitycloudservice/srinidhi.koushik@oracle.com',
                'oracleidentitycloudservice/partha.mahanta@oracle.com'
                ]
    '''
    
    vlistname=['iaas.us-ashburn-1','ap-hyderabad-1']
    for i in vlistname:
        print(i)
        nm = i.split("-")   # splits into two words as oracleidentitycloudservice  srinidhi.koushik@oracle.com
        print(nm[1])        # srinidhi.koushik@oracle.com
                
def risetimes():
    times = []

    for rt in risetimes:
        time = datetime.fromtimestamp(rt)
        times.append(time)
        print(time)

def daydiff():
    #dt1 = '20-11-24T04:26:46'

    

    now = datetime.now()    
    t = now.strftime("%H:%M:%S")
    print("time:", t,type(t))           # 14:59:38

    s1 = now.strftime("%m/%d/%Y, %H:%M:%S")
    # mm/dd/YY H:M:S format
    print("s1:", s1,type(s1))           # 07/13/2021, 14:59:38

    s2 = now.strftime("%d/%m/%Y, %H:%M:%S")
    # dd/mm/YY H:M:S format
    print("s2:", s2,type(s2))           # 13/07/2021, 14:59:38

    s3 = now.strftime("%H:%M:%S")
    print("s3:", s3,type(s3))           # 13/07/2021, 14:59:38


    print(" Perfectly working str to time for time difference  ")
    t1 = '14:55:22'
    t2 = '15:03:38'    
    start_time = datetime.strptime(t1, '%H:%M:%S')
    #print(start_time,type(start_time))
    end_time = datetime.strptime(t2, '%H:%M:%S')
    #print(end_time,type(end_time))
    time_diff = end_time - start_time
    time_min = int(time_diff.seconds / 60)
    time_sec = int(time_diff.seconds % 60)
    time_taken = "%s min %s sec" %(time_min, time_sec)
    print(time_taken)


    print('Perfectly working  current time with old time for time difference')
    t1 = datetime.now()
    #print(t1,type(t1))
    t11 = t1.strftime("%H:%M:%S")
    time.sleep(5)
    t2 = datetime.now()
    #print(t2,type(t2))
    t21 = t2.strftime("%H:%M:%S")

    start_time = datetime.strptime(t11, '%H:%M:%S')
    #print(start_time,type(start_time))
    end_time = datetime.strptime(t21, '%H:%M:%S')
    #print(end_time,type(end_time))

    #end_time = t2 - t1
    time_diff = end_time - start_time
    time_min = int(time_diff.seconds / 60)
    time_sec = int(time_diff.seconds % 60)
    time_taken = "%s min %s sec" %(time_min, time_sec)
    print(time_taken)



    print("Perfectly working str to Date for Days difference.. ")
    dt1 = '21-11-24T'
    dt2 = date.today()
    #date_format = "%m/%d/%Y"
    date_format = "%Y/%m/%d"
    a = datetime.strptime(dt1, '%y-%m-%dT').date()
    #print(type(a),(a))
    #print(dt2,type(dt2))
    delta = a - dt2
    print(delta) #  Number of days
    

def writefile():
    flobj = open(r"C:\MyFolder\PyLearn\ToRef\hostname-0608.txt","r")
    flobj.readlines
    #rgstr = "None"
    cluname = "None"
    regncode = "None"
    regnlst = ['hyd']
    #regnlst = [rgnstr]
    #print(type(regnlst))
    #print(regnlst)
    mdict = {}
    cdict = {}
    #jsondata = []
    Flg = 'FALSE'
    #for lst in regnlst:
    #    print(lst)
    for line in flobj:
        vline = line.strip()
        #print(regnlst)
        '''
        if any(c in vline for c in regnlst):
            print("line>>",vline,"regnlst >>",regnlst)
            cluname = line[1:44]
            regncode = line[1:4]          #hyd
            
        #print(line)
        '''
        vline = line.strip()
        if vline.startswith('(') and vline.endswith(')'):
            print("startswith")
            #if regnlst in vline:
            #if lst in vline:
            #    cluname = vline[1:44]
            #    regncode = vline[1:4]          #hyd
            if any(c in vline for c in regnlst):
                print("line>>",vline,"regnlst >>",regnlst)
                cluname = line[1:44]
                regncode = line[1:4]          #hyd
                
                    
        if 'C:\Program Files' in vline:
            constr = vline[46:188]
            if regncode in constr:
                Flg='TRUE'
                #print("regncode >> ",regncode,"clustname>> ",cluname)
                x_pie = vline.split('"')
                #conStr = "'" + '"'+ x_pie[1] + '"' + x_pie[2] + '"' + x_pie[3] + '"' + x_pie[4] + '"' + x_pie[5] + '"' + "'"
                print("x_pie-1",x_pie[1])
                print("x_pie-2",x_pie[2])                
                print("x_pie-3",x_pie[3])                
                print("x_pie-4",x_pie[4])                
                print("x_pie-5",x_pie[5])                
                
                #acd_api.py -> atp_oci_acd_adp_dict = {"ACD":acd_ocid_list,"ADB":[]}
                
                pstr="C:\Program Files (x86)\PuTTY\plink.exe"
                #print("pstr :",pstr)
                cdict.update({cluname : x_pie[2]+x_pie[3]+" "+x_pie[4].strip()})
                
                print("cdict >>>>>" , cdict)
                print("\n")
                #body = json.dumps({regncode:{clunm:x_pie[2]+x_pie[3]+x_pie[4]})
                #print(body)
                #print(json.dumps({"regn":rgstr,"connectstr":consstr}))
                #jsondata_list.append(body)
                
                #mdict.update({regncode:cdict})
                #print("mdict inside loop >>> is ",mdict)
                #print("\n")
            #print(json.dumps(mdict))
        
    if Flg == 'TRUE':
        #mdict.update({regncode:cdict})        
        #print("mdict outside >>> is ",mdict)     
        #print("\n")
        print("jsondump>>> ",json.dumps(cdict))    
        #with open('src.json', 'a') as fp:
        #    json.dump(cdict,fp)
        #    fp.close()
                        
def readfile():
    #in_dod is dict of dict as below
    print('entering')
    in_dod =   {"hyd":{"hyd100402exd-d0-03-04-cl-04-06-clu01-n9w1i1":"hyd1connestr",
                "hyd200402exd-d0-03-04-cl-04-06-clu01-n9w1i2":"hyd2connestr",
                "hyd300402exd-d0-02-04-cl-04-06-clu01-m9w1i1":"hyd3connestr",
                "hyd400402exd-d0-01-04-cl-04-06-clu01-m9w1i2":"hyd4connestr"},
                "zrh":{"zrh100615exd-d0-05-06-cl-07-09-clu01-ac4pm1":"zrh1connectstr",      
                "zrh200615exd-d0-05-06-cl-07-09-clu01-ac4pm2":"zrh2connectstr",
                "zrh300615exd-d0-03-06-cl-07-09-clu01-ac4pm3":"zrh3connectstr",
                "zrh400615exd-d0-03-06-cl-07-09-clu01-ac4pm4":"zrh4connectstr",
                "zrh500615exd-d0-02-06-cl-07-09-clu01-ac4pm5":"zrh5connectstr"}
                }
   
    #in_dod =json.load(open(r"C:\MyFolder\PyLearn\ToRef\src.json","r"),object_pairs_hook=OrderedDict)   
    
    #print("input_dict ",in_dod)
    #print("\n")    
    #print("input_dict hyd ",in_dod["hyd"])
    #print("\n")    
    #print("input_dict zrh ",in_dod["zrh"])
    #print("\n")    
            
     
    d2 = {}
    print(in_dod.keys())
    values = in_dod.keys()
    print(values)
    values_list = list(values)
    print(values_list)
    cnt = 0
    for k, v  in in_dod.items():
        #print("k type is",type(k))
        #print("v type is",type(v))
        if 'fra' in k:
            print("key >>",k, '-> values are ',v)
        #if k == 'zrh' :
        
        cnt = cnt + 1
        print('cnt is >>>> ', cnt)
        d2 = v
        for k1,v1 in d2.items():
            print("key1 >>",k1, '-> values are ',v1)
                
    '''
    #https://www.geeksforgeeks.org/iterate-over-a-dictionary-in-python/
    in_dod1 =  {"hyd":{"clu":"hyd100402exd-d0-03-04-cl-04-06-clu01-n9w1i1","cstr":"hyd1connestr",
                "clu":"hyd200402exd-d0-03-04-cl-04-06-clu01-n9w1i2","cstr":"hyd2connestr",
                "clu":"hyd300402exd-d0-02-04-cl-04-06-clu01-m9w1i1","cstr":"hyd3connestr",
                "clu":"hyd400402exd-d0-01-04-cl-04-06-clu01-m9w1i2","cstr":"hyd4connestr"},
                "zrh":{"clu":"zrh100615exd-d0-05-06-cl-07-09-clu01-ac4pm1","cstr":"zrh1connectstr",      
                "clu":"zrh200615exd-d0-05-06-cl-07-09-clu01-ac4pm2","cstr":"zrh2connectstr",
                "clu":"zrh300615exd-d0-03-06-cl-07-09-clu01-ac4pm3","cstr":"zrh3connectstr",
                "clu":"zrh400615exd-d0-03-06-cl-07-09-clu01-ac4pm4","cstr":"zrh4connectstr",
                "clu":"zrh500615exd-d0-02-06-cl-07-09-clu01-ac4pm5","cstr":"zrh5connectstr"}
                }
               
               
    print(type(in_dod1))            
    print("input_dict ",in_dod1)
    print("\n")    
    print("input_dict hyd ",in_dod1["hyd"])
    print("input_dict cluster",in_dod1["hyd"]["clu"])
    print("input_dict clus with str",in_dod1["hyd"]["cstr"])
    print("\n")    
    print("input_dict zrh ",in_dod1["zrh"])
    print("input_dict cluster",in_dod1["zrh"]["clu"])
    print("input_dict clus with str",in_dod1["zrh"]["cstr"])
    
    print(in_dod1.values())

    print("\n")    
    d2 = {}
    for k, v  in in_dod1.items():
        print("k type",type(k))
        print("v type",type(v))
        print("key >>",k, '-> values are ',v)
        if k == 'zrh' :
            d2 = v
            for k1,v1 in d2.items():
                print("key1 >>",k1, '-> values are ',v1)
            
    for key in in_dod:
        print("key is>>", key ,">>",in_dod[key])
        #print("cluster are >>",in_dod["cluster"])
        #print("str are >>",in_dod["str"])
        
    print("\n")    
    
    
    for val in in_dod.values():
        print("Values are>>",val)
        #print("cluster are >>",in_dod["cluster"])
        #print("str are >>",in_dod["str"])     
    
    print("\n")            
    #input_dict=json.load(open(r"C:\MyFolder\PyLearn\ToRef\dod.json","r"), object_pairs_hook=OrderedDict)
    #input_dict=json.load(open(r"C:\MyFolder\PyLearn\ToRef\dod.json","r"))
    #input_dict=json.load(open(r"C:\MyFolder\PyLearn\ToRef\cmd.json","r"))

    

        #print("cluster are >>",input_dict["cluster"])
        #print("str are >>",input_dict["str"])
    
    
    for key in in_dod:
        print("key is>>", key ,">>",in_dod[key])
        #print("cluster are >>",in_dod["cluster"])
        #print("str are >>",in_dod["str"])
        
    print("\n")    
    
    
    for val in in_dod.values():
        print("Values are>>",val)
        #print("cluster are >>",in_dod["cluster"])
        #print("str are >>",in_dod["str"])
    
    print("\n")            

    
def readjson():
mf-fund.json 
        {"Aditya Birla SunLife ESG Fund Reg":"20",	
        "Aditya Birla SunLife Multi-Cap Fund Reg":"30",	
        "DSP Flexi Cap Fund":"20",						
        "HDFC Retirement Savings Fund Equity Reg":"20",	
        "IDFC Emerging Businesses Fund Reg":"20",			
        "IDFC Large Cap Reg":"20",						
        "IDFC Multi Cap Fund Reg":"20",
        "Kotak Emerging Equity Reg":"20",
        "Kotak Equity Opportunities Reg":"20",
        "Nippon India Retirement Fund - Wealth Creation Scheme":"20",
        "Sundaram Balanced Advantage Fund Reg":"20",
        "Sundaram Equity Fund Reg":"20",
        "Tata Business Cycle Fund Reg":"20",
        "Tata Large & Mid Cap Reg":"20",
        "Tata Multi Asset Opportunities Fund Reg-G":"20",
        "HDFC Children's Gift Fund-Inv":"20"}
        
getvaluefor  = read_input_json('mf-fund.json')
    #for k,v in getvalue.items():
    #  print(k, v)
    #print(getvaluefor.keys())
    #print(getvaluefor.values())
'''    
def readdictlist():

    in_dol = {
            "ROOT" : ["/home/achanpur/demo_ready/backup_triage_jun_1_0 izwwl6vz -e"],
            "GI" :   [],		
            "DB" : []
             }

    #content=json.load(open(ifile))        
    print(type(in_dol))
    #tmpstr = ''
    cmdstr = ''
    endstr = "'"
    for k in in_dol.keys():
        print("K is >>",k)
        cmdlist = in_dol[k]
        # tmpstr = constr(cmdlist)
        tmpstr = ''
        #print("cmdlist >> ",cmdlist)
        #print(type(cmdlist))      # list
        if k == 'ROOT' and ( len(cmdlist) > 0): 
            sudostr = "sudo bash -c '"
            tmpstr=sudostr
        elif k == 'GI' and ( len(cmdlist) > 0): 
            sudostr = "sudo su - grid -c '"        
            tmpstr=sudostr
        elif k == 'DB' and ( len(cmdlist) > 0): 
            sudostr = "sudo su - oracle -c '"            
            tmpstr=sudostr
            
        lste = len(cmdlist)-1

        for lines in range(0,len(cmdlist)):
            lstr = cmdlist[lines]
            if lste != int(str(lines)):
                tmpstr = tmpstr + lstr + ";"
                #print("tmpstr >>>>",tmpstr)
            elif lste == int(str(lines)):
                #print("is last element TRUE >>>>>>> : ", lste,str(lines))
                tmpstr = tmpstr + lstr + "';"
            
        cmdstr = cmdstr + tmpstr          
    print("cmdstr >>>> ",cmdstr,'\n')
    
    return cmdstr + '"'
    #strSudo = cmdstr + '"'

def readmissingkey_indict():
    dict1 = {'Name': 'Zara', 'Age': 27}
    dict2 = {"aa": "aaa", "definedTags": {"default_tags": "yes"}}
    
    d =   {"a":
            {"aa": "aaa", 
                "definedTags": {"default_tags": {"CreatedBy": "raji.com",  "CreatedOn": "021"}},
            'Name': 'Zara','Age': 27},
            "b":{"bb": "bbb", 
            "definedTags": {"Oracle-Tags": {"CreatedBy": "crsk.com",  "CreatedOn": "022"}},
            'Name': 'Thara', 'Age': 37}
            }
            
    d1 = [{"a":
            {"aa": "aaa", 
                "definedTags": {"default_tags": {"CreatedBy": "raji.com",  "CreatedOn": "021"}},
            'Name': 'Zara','Age': 27},
            "b":{"bb": "bbb", 
            "definedTags": {"Oracle-Tags": {"CreatedBy": "crsk.com",  "CreatedOn": "022"}},
            'Name': 'Thara', 'Age': 37}
            }]
            
    '''    
    dict = {
            {"ad": "bxt", "definedTags": {"default_tags": {"CreatedBy": "e.com",  "CreatedOn": "021"}},
            "ad": "bxt2", "definedTags": {"Oracle-Tags": {"CreatedBy": "e.com",  "CreatedOn": "021"}}}
            }
    '''       
    #print ("Value : %s" %  d.get([0][0]['definedTags'], "NA"))
    #print ("Value : %s" %  d.get([1]['definedTags'], "NA"))
    print(type(d),len(d))
    
    tmpd = {}
    for i in d.items():
        print((i),len(i),"-",print(type(i)))
        print('i[0] >',i[0])        # a
        #print('i[0] >',i[0]['definedTags'])
        print('i[1] >',i[1])        # aa.. till 27
        print('i[1] definedTags >',i[1]['definedTags']) # from default_tag..till 021
        print('i[1] >',i[1]['definedTags']) 
        tmpd = i[1]['definedTags']
        print('tmpd',type(tmpd),tmpd)
        
        print ("Value : %s" %  tmpd.get('default_tags'))
        dcrby = tmpd.get('default_tags')
        if dcrby == 'NONE':
            print(vcrby['CreatedBy'])
        
        print ("Value : %s" %  tmpd.get('Oracle-Tags'))
        ocrby = tmpd.get('Oracle-Tags')
        if ocrby == 'NONE':
            print(vcrby['CreatedBy'])


    '''        
        print ("Value : %s" %  i.get('ad',"NA"))
        print ("Value : %s" %  i.get('definedTags',"NA"))
        print ("Value : %s" %  i.get('Oracle-Tags',"NA"))
        print ("Value : %s" %  dict.get('definedTags', "NA"))        
    '''        
        


'''
Date: 030721
Correct: json format.  q2a: Try to read this 

createEmployeeDictionary = {
       "main_key": {            <<<<<<  correct
            "PersonNumber": "x",
            "names": [
                {
                    "LastName": "x",
                    "FirstName": "x",
                    "LegislationCode": "US"
                }
            ],
            "workRelationships": [
                {
                    "LegalEmployerName": "ZBEN USA",
                    "WorkerType": "E",
                    "PrimaryFlag": True,
                    "StartDate": "2021-06-17",
                    "assignments": [
                        {
                            "BusinessUnitName": "ZBEN_Common_Business Unit",
                            "ActionCode": "HIRE",
                            "UserPersonType": "Employee",
                            "AssignmentNumber": "x",
                            "managers": [
                                {
                                    "ManagerAssignmentNumber": "y",
                                    "ManagerType": "LINE MANAGER"
                                }
                            ]
                        }
                    ]
                }
            ]
        }
    }
    
Wrong:
createEmployeeDictionary = {
        {
            "PersonNumber": "x",
            "names": [
                {
                    "LastName": "x",
                    "FirstName": "x",
                    "LegislationCode": "US"
                }
            ],
            "workRelationships": [
                {
                    "LegalEmployerName": "ZBEN USA",
                    "WorkerType": "E",
                    "PrimaryFlag": True,
                    "StartDate": "2021-06-17",
                    "assignments": [
                        {
                            "BusinessUnitName": "ZBEN_Common_Business Unit",
                            "ActionCode": "HIRE",
                            "UserPersonType": "Employee",
                            "AssignmentNumber": "x",
                            "managers": [
                                {
                                    "ManagerAssignmentNumber": "y",
                                    "ManagerType": "LINE MANAGER"
                                }
                            ]
                        }
                    ]
                }
            ]
        }
    }    
'''    
    


def numpylist():
    in_value=[1,2,3,4,5,6,7,8,9]
    input_value = np.array(in_value)
    print(str(input_value.reshape((3,3)))[1:-1])
    
    

    

    
    

def pause():
    programPause = raw_input("Press the <ENTER> key to continue...")
        



    #return input_dict
   

def sendemail():
#https://realpython.com/python-send-email/#option-2-setting-up-a-local-smtp-server
    try:
        sender = 'sathishkumar.rangaraj@oracle.com'
        receivers = 'rashmi.panda@oracle.com'

        message = """From: From Person <from@fromdomain.com>
            To: To Person <to@todomain.com>
            Subject: SMTP e-mail test
        
            This is a test e-mail message.
            """


        smtpObj = smtplib.SMTP('localhost')
        smtpObj.sendmail(sender, receivers, message)         
        print("Successfully sent email")
   
        #except SMTPException:
    except Exception as err :
        print("Error: unable to send email",err)
   
def sendemailg():

#https://realpython.com/python-send-email/#code-example
    '''
    Oracle policy is that Exchange is only used for interactive email clients, and that other mail routers are used for scripts, 
    such as internal-mail-router.oracle.com for things behind the Oracle firewall.  
    https://confluence.oraclecorp.com/confluence/display/OITGLOBAL/Sendmail+SMTP+Server+Configuration+Procedure
    
    message = """Subject: Your grade

    Hi {name}, your grade is {grade}"""
    sender = "4crsksocial@gmail.com"
    receivers = 'sathishkumar.rangaraj@oracle.com'

    port = 465  # For SSL
    #password = input("Type your password and press enter: ")
    password = 'Mypwd2275#'
    # Create a secure SSL context
    context = ssl.create_default_context()

    with smtplib.SMTP_SSL("smtp.gmail.com", port, context=context) as server:
        server.login("4crsksocial@gmail.com", password)
        # TODO: Send email here
        server.sendmail(from_address,receivers,message)
    
    Variable: HTTP_PROXY
    Value: http://proxyserver:port and
    Variable: HTTPS_PROXY
    Value: https://proxyserver:port
    
    set http_proxy=http://[username:password@]proxyserver:port
    set https_proxy=https://[username:password@]proxyserver:port

    This is what works for me with VPN.. py -m pip install --proxy www-proxy.us.oracle.com numpy        
        
    '''        
    # prerna.bhise@oracle.com   <<<<< contact person
    # http://vpad/vpad.dat
    # http://wpad.idc.oracle.com/wpad.dat
    # https://artifactory.oci.oraclecorp.com/list/ops-proxyconf-local/proxy.pac
    # www-proxy.us.oracle.com
    sender_email = "sathishkumar.rangaraj@oracle.com"   #you can create common email id
    smtp_server = "internal-mail-router.oracle.com"
    smtp_port = 25
    message = MIMEMultipart("alternative")
    message["Subject"] = "Test Email 2 from python"
    message["From"] = "sathishkumar.rangaraj@oracle.com"  #you can use sender_email value
    message["To"] = ''
    addresses = ['sathishkumar.rangaraj@oracle.com','nalini.m@oracle.com']
    
    for address in addresses:
        if message["To"] == '':
            message["To"] = address
        else:
            message["To"] = message["To"] + "," + address
            
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        for address in addresses:
            server.sendmail(
                sender_email, address, message.as_string()
            )
        server.quit()    

def csv2json():

    '''
    https://realpython.com/python-csv/
    https://www.marsja.se/your-guide-to-reading-excel-xlsx-files-in-python/
    https://www.geeksforgeeks.org/reading-excel-file-using-python/
    https://linuxhint.com/read-excel-file-python/
    https://www.codegrepper.com/code-examples/python/how+to+read+xlsx+file+in+python
    
    Updated in csv(excel)
    testdatacsv.csv
    oggdeployDetails region     iaas.eu-frankfurt-1
                     apiversion 20160918
                     pdbocid    ocid1.autonomo
    targetApiDetails region     iaas.eu-fra
                     walletpassword Walletpwd,123   << one word along with ',' 
                     name       sathishkumar
                     age        45

    stored as testdatacsv.csv
    oggdeployDetails,region,iaas.eu-frankfurt-1
    ,apiversion,20160918
    ,pdbocid,ocid1.autonomo
    targetApiDetails,region,iaas.eu-fra
    ,walletpassword,"Walletpwd,123"
    ,name,sathishkumar
    ,age,45


    '''
    vdict={}
    vdict2={}
    col0=''
    with open('testdatacsv.csv','r') as csvf:
        csvreader = csv.reader(csvf,delimiter = ",", quotechar='"') # "\t"
        for row in csvreader:
            #print(row)
            #print('colums are {", ".join(row)}')
            print("co10: ",row[0],"co11: ", row[1],"co12: ",row[2])
            vdict.update({row[1]:row[2]})  # ok
            
            #vdict.update({row[0]:row[1]+':'+row[2]}) 
                # updates as {'oggdeployDetails': 'pdbocid:ocid1.autonomo', 'targetApiDetails': 'age:45'}
            #vdict.update({row[0][1]:row[2]})
                #LL: it updates as {'g': 'ocid1.autonomo', 'a': '45'}
                # not as dict of dict
            '''        
            col0=row[0]
            if vcol == col0:                
                vdict.update({row[0]:{row[1]:row[2]}}) 
                    # updates as {'oggdeployDetails': {'pdbocid': 'ocid1.autonomo'}, 'targetApiDetails': {'age': '45'}}
            elif vcol != col10:
                vdict1.update(vdict)
                
            col0=''
            vcol=row[0]
            '''
            
            print(vdict)

    with open('teststoreasjson.json','a') as fp:    
        json.dump(vdict,fp)
        fp.close
        
    '''
        teststoreasjson.json stored as 
            {"region": "iaas.eu-fra", "apiversion": "20160918", 
             "pdbocid": "ocid1.autonomo", "walletpassword": "Walletpwd,123", 
             "name": "sathishkumar", "age": "45"}
    '''  

def readcsvjson():
    in_dod = json.load(open("teststoreasjson.json","r"))
    for k, v  in in_dod.items():
        print(k ,'-',v)
        
def csvdictreader():
    vdict={}
    with open('testdatacsv.csv','r') as csvf:
        csvreader = csv.DictReader(csvf,delimiter = ",", quotechar='"') # "\t"
        for row in csvreader:
            print({row["oggdeployDetails"]},{row["region"]})
            #LL: it reads value in column format.. i.e only col(0),col(1)
            
            #vdict.update({row[1]:row[2]})
            #vdict.update({row[0]:row[1]+':'+row[2]})
           

def xls2json():
    vdict = {}
    wb = openpyxl.load_workbook("userdata_test.xlsx")
    ws = wb['UserData']
    print(ws)
    print("number of rows",ws.max_row,"number of cols ",ws.max_column)

    '''
    for row in ws.iter_rows(max_row=8):
        for cell in row:
            print(cell.value,end=" ")
    '''
    #for row in range(0,ws.max_row):
    for row in ws.iter_rows(max_row=ws.max_row):
        #for col in ws.iter_cols(2,ws.max_column):
        #for col in ws.iter_cols(2,3):
        print(row[1].value, row[2].value)
        vdict.update({row[1].value : row[2].value})  # ok
        '''
        for col in range(1,2):
            print("rows >",row)
            print("cols >",col)
            #print(row[col][0].value,row[col][1].value)
            #print("co11: ", col[0].value, "co12: ", col[1].value)
            #print(row[col].value)
            #print(col[row].value,end="\t\t")
        '''

    print(vdict)

    with open('ogg_userdata_test.json','w') as fp:
        json.dump(vdict,fp)
        fp.close
    wb.close()    


def read_input_json(file_name):
    #input_dict=json.load(open(file_name) % {"src_ggadmin_user_name":src_ggadmin_user_name_v})
    #{"src_net_alias_name":src_net_alias_name_v}{"src_schema_toreplicate":src_schema_toreplicate_v}) 
    #"ct=$(date '+%Y%m%d_%H%M%S'); mv /home/opc/tmp/addschematran.oby /home/opc/tmp/addschematran.oby.$ct",    
    dtfmt_v='+%Y%m%d_%H%M%S'
    print(type(src_ggadmin_user_name_v),' - ',src_ggadmin_user_name_v)
    with open(file_name) as f:
        input_dict = json.loads(f.read() % {'src_ggadmin_user_name': src_ggadmin_user_name_v , \
        'src_ggadmin_user_pwd': src_ggadmin_user_pwd_v , \
        'src_ddl_user_name': src_ddl_user_name_v , \
        'src_ddl_user_pwd': src_ddl_user_pwd_v , \
        'src_ddl_user_tbs_name': src_ddl_user_tbs_name_v , \
        'src_ddl_user_tbs_df': src_ddl_user_tbs_df_v , \
        'src_schema_toreplicate': src_schema_toreplicate_v , \
        'src_schema_toreplicate_pwd': src_schema_toreplicate_pwd_v , \
        'src_table_toreplicate': src_table_toreplicate_v , \
        'src_tns_name' : src_tns_name_v , \
        'src_host_port' : src_host_port_v , \
        'src_host_ip': src_host_ip_v , \
        'src_service_name' : src_service_name_v , \
        'tgt_ggadmin_user_name' :tgt_ggadmin_user_name_v , \
        'tgt_tns_name':tgt_tns_name_v , \
        'tgt_ggadmin_user_pwd':tgt_ggadmin_user_pwd_v , \
        'tgt_schema_tomap':tgt_schema_tomap_v , \
        'dtfmt': dtfmt_v}
        )  

    return input_dict       
    
def read_input(file_name):
    input_dict=json.load(open(file_name))
    return input_dict



def readjsonwitharg(file_name):
# https://stackoverflow.com/questions/62195181/how-to-pass-variable-to-json-for-python
       #how to pass variable as value in json python
    #getvaluefor = read_input_json('ogg_input_test.json')   
    getvaluefor = read_input_json(file_name)   
    commands = getvaluefor['gg_create_ext_prm']
    for cmdl in commands:
        print(cmdl)
                
    commands = getvaluefor['gg_add_schematrandata']
    for cmdl in commands:
        print(cmdl)
    
    print('fortnssrc')
    commands = getvaluefor['tnssrc']
    for cmdl in commands:
        print(cmdl)
        
    print(' connection_info ')        
    commands = getvaluefor['connection_info']
    for k, v  in commands.items():
        print( k , ' - ', v)
        
    print(' gg_create_ddlsh ')        
    commands = getvaluefor['gg_create_ddlsh']
    for cmdl in commands:
        print(cmdl)
        
    print(' gg_create_ddlsh_org ')                
    commands = getvaluefor['gg_create_ddlsh_org']
    for cmdl in commands:
        print(cmdl)
    
    
    #with open("path/to/json") as f:
    #    rendered = json.loads(f.read() % {"user_name": username})
    
    
    #{"path_to_folder": "/Users/%(user_name)s/my_folder/"}
   
   
def txt2xls():
    #https://www.pythonexcel.com/openpyxl-write-to-cell.php
    vdict = {}
    wb = openpyxl.load_workbook('mfdata.xlsx')
    userdata_ws = wb['Sheet1']
    i = 0

    flobj = open(r"SrcData.txt","r")
    flobj.readlines
    for line in flobj:
        vline = line.strip()
        #print(vline)
        mfname = vline.split("-G")
        print(mfname, '- ', len(mfname))
        if len(mfname) > 1:
            mfnm = mfname[0]
            mfvalues = mfname[1]
            #print('mfname > ',mfname)
            #print('mfname[1]>',mfvalues)
            nm = mfvalues.split()   # splits into two words as oracleidentitycloudservice  srinidhi.koushik@oracle.com
            print(nm, ' len is ' , len(nm))

            vFolio = nm[0]
            vStDate = nm[1]
            vUnits = nm[2]
            vAvg   = nm[3]
            vCP     = nm[4]
            vNAV    = nm[5]
            vMV     = nm[6]
            vAbsRet = nm[8]
            # nm[7] PortWt
            #vXRR    = nm[9]
            i = i + 1
            '''
            vFolioC =  userdata_ws.cell(row=i, column=1)
            vStDateC =  userdata_ws.cell(row=i, column=2)
            vUnitsC =  userdata_ws.cell(row=i, column=3)
            vAvgC =  userdata_ws.cell(row=i, column=4)
            vCPC =  userdata_ws.cell(row=i, column=5)
            vNAVC =  userdata_ws.cell(row=i, column=6)
            vAbsRetC =  userdata_ws.cell(row=i, column=7)
            vFolioC =  userdata_ws.cell(row=i, column=8)
            vFolioC.value = vFolioC
            vStDateC.value = vStDate
            vUnitsC.value = vUnits
            vAvgC.value =  vAvg
            vCPC.value =  vCP
            vNAVC.value =  vNAV
            vAbsRetC.value = vAbsRet
            '''
            userdata_ws.cell(row=i, column=1).value = mfnm
            userdata_ws.cell(row=i, column=2).value = vFolio
            userdata_ws.cell(row=i, column=3).value = vStDate
            userdata_ws.cell(row=i, column=4).value = vUnits
            userdata_ws.cell(row=i, column=5).value = vAvg
            userdata_ws.cell(row=i, column=6).value = vNAV
            userdata_ws.cell(row=i, column=7).value = vMV
            userdata_ws.cell(row=i, column=8).value = vAbsRet

            
            print(' ')
        mfname = vline.split("-in)")
        print(mfname, '- ', len(mfname))            
        if len(mfname) > 1:
            mfvalues = mfname[1]
            #print('mfname > ',mfname)
            #print('mfname[1]>',mfvalues)
            nm = mfvalues.split()   # splits into two words as oracleidentitycloudservice  srinidhi.koushik@oracle.com
            print(nm, ' len is ' , len(nm))
            vFolio = nm[0]
            vStDate = nm[1]
            vUnits = nm[2]
            vAvg   = nm[3]
            vCP     = nm[4]
            vNAV    = nm[5]
            vMV     = nm[6]
            vAbsRet = nm[8]
            print(' ')     
            i = i + 1


            
        #for vMF in nm:
        #    print(vMF)  
        
        
        print('vFolio',vFolio) 
        print('vStDate',vStDate)
        print('vUnits', vUnits)
        print('vAvg', vAvg)  
        print('vCP',vCP)    
        print('vNAV',vNAV)   
        print('vMV',vMV)    
        print('vAbsRet',vAbsRet)

        print(' ')
        
   
    wb.save('mfdata.xlsx')    
    '''
         
    >>>mycell= mysheet.cell(row=4, column=5)    
    >>>mycell='Writing data to E4'
    Using a loop to write to a series of cells picking values from a list.
    >>>for i in range(5,15):
            cellref=mysheet.cell(row=i, column=5)
            cellref.value=lista[i]
           
            
    for row in userdata_ws.iter_rows(max_row=userdata_ws.max_row):
    #for row in userdata_ws.iter_rows(max_row=26):
        #print(row[1].value, row[2].value)
        vdict.update({row[1].value : row[2].value})  # ok

    for row in ocidet_ws.iter_rows(max_row=ocidet_ws.max_row):
    #for row in ocidet_ws.iter_rows(max_row=15):
        #print(row[1].value, row[2].value)
        vdict.update({row[1].value : row[2].value})  # ok


    #print(vdict)
    print("Scanned ", wb.sheetnames[0],userdata_ws.max_row," : rows Updated in config file")
    print("Scanned ", wb.sheetnames[1],ocidet_ws.max_row," : of rows Updated in config file")

        if vline.startswith('(') and vline.endswith(')'):
            print("startswith")
            #if regnlst in vline:
            #if lst in vline:
            #    cluname = vline[1:44]
            #    regncode = vline[1:4]          #hyd
            if any(c in vline for c in regnlst):
                print("line>>",vline,"regnlst >>",regnlst)
                cluname = line[1:44]
                regncode = line[1:4]          #hyd
    '''  

    

#sehubdb_high = (description= (retry_count=20)(retry_delay=3)(address=(protocol=tcps)(port=1522)(host=adb.ap-mumbai-1.oraclecloud.com))
#	(connect_data=(service_name=diwqiplghkb2x0o_sehubdb_high.adb.oraclecloud.com))
#	(security=(ssl_server_cert_dn="CN=adb.ap-mumbai-1.oraclecloud.com, OU=Oracle ADB INDIA, O=Oracle Corporation, L=Redwood City, ST=California, C=US")))
#"check_dblogmode":{"select log_mode from v$database":"ARCHIVELOG"},


def srcdbconnect():
    vsrcdbIP = "NA"
    vPort = 1522
    vSrvcName = 'diwqiplghkb2x0o_sehubdb_high.adb.oraclecloud.com'
    vdbauser = 'admin'
    vdbapwd = 'Welcome__pwd123'

    #logmain.info('connecting to Source database')
    #dsn_tns = cx_Oracle.makedsn(vsrcdbIP,vPort, service_name=vSrvcName) 
    sehubdb_high = '(description= (retry_count=20)(retry_delay=3)(address=(protocol=tcps)(port=1522)(host=adb.ap-mumbai-1.oraclecloud.com))	(connect_data=(service_name=diwqiplghkb2x0o_sehubdb_high.adb.oraclecloud.com))	(security=(ssl_server_cert_dn="CN=adb.ap-mumbai-1.oraclecloud.com, OU=Oracle ADB INDIA, O=Oracle Corporation, L=Redwood City, ST=California, C=US")))'

    #print(dsn_tns)
    #conn = cx_Oracle.connect(user=vdbauser, password=vdbapwd, dsn=sehubdb_high,mode = cx_Oracle) 
    conn = cx_Oracle.connect('admin/Welcome__pwd123@sehubdb_high')
    print(conn.version)
    #conn = cx_Oracle.connect(user=vdbauser, password=vdbapwd, dsn=dsn_tns) 
    return conn

    
def fn_srcdb_check_dblogmode(vCurobj):
    # https://www.geeksforgeeks.org/oracle-database-connection-in-python/
    # https://www.oracletutorial.com/python-oracle/
    # https://www.oracletutorial.com/python-oracle/connecting-to-oracle-database-in-python/
    
    vcurr = vCurobj
    vowner_nm = 'JAPAC_TENANCY'
    Vtab_na = 'Peter'
    try:
        #vcurr.execute("create table employee(empid integer primary key, name varchar2(30), salary number(10, 2))")

        #qry = ('select ' + 
        #        'log_mode ' + 
        #        'from v$database')
        #print(qry)
        #vSql = """select owner,table_name from dba_tables where owner =  :own_nm """
        #vcurr.execute(vSql, own_nm = vowner_nm)
        
        #vcurr.execute('insert into employee values(10001,\'Rahul\',50000.50)')

        #data = [[10007, 'Vikram', 48000.0], [10008, 'Sunil', 65000.1], [10009, 'Sameer', 75000.0]]
        #vcurr.executemany('insert into employee values(:1,:2,:3)', data)
        
        #conn.commit()
        '''
        vcurr.execute(qry)
        row_value = vcurr.fetchall()  # [0]
        for i in row_value:
            #if row_value is None:
            print(i)
        '''
        
        dbv_instancename = 'crsk_instance'
        dbv_vCreatedBy = 'sathi.rang@oracle.com'
        vSchema_nm = 'JAPAC_TENANCY'
        vTable_nm = 't_cloud_usage'
        dbv_createdon = '021-09-13T04:54:30'
        
        #data = [dbv_displayname,dbv_vCreatedBy]
        print(dbv_instancename,dbv_vCreatedBy,dbv_createdon[0:9])

        data = [dbv_instancename,dbv_vCreatedBy]
        print(data)
        
        vSql = ('insert into ' + vSchema_nm + '.' + vTable_nm + 
                '(instance_name,se_name,CREATED_ON) values ' +
                '(\'' + dbv_instancename + '\',' + 
                '\'' + dbv_vCreatedBy + '\',' +
                'to_date(\'' + dbv_createdon[0:9] + '\',\'YYYY/MM/DD\'))')
        print(vSql)
        vcurr.execute(vSql)
        conn.commit()
    
     #insert into JAPAC_TENANCY.t_cloud_usage(instance_name,se_name,CREATED_ON) values 
     #   ('MysqlClient_Mumbai','sathishkumar.rangaraj@oracle.com',to_date('021-09-13','YYYY/MM/DD'));
     
    except cx_Oracle.DatabaseError as err:
        print("There is a problem with Oracle", err)

    finally:
        if vcurr:
            vcurr.close()
            print('cursor closed')
        if conn:
            conn.close()
            print('connection closed')
        


    
    
if __name__=="__main__":
    #writefile() 
    #writefile('zrh') 
    #print("Think about what you ate for dinner last night...")
    #pause()
    #readfile()
    
    #numpylist()
    
        
        
        #print('i[1] >',i[1])
        #print('i[2] >',i[2])
        
        
    
    #daydiff()  # good
    #sendemail()
    #sendemailg() 
    #splitword() # ok
    #printalign()  # ok  
    #spool2text()    
    
    # print(os.path.abspath(os.getcwd()))    # to print current path
    # print(os.path.dirname(os.path.abspath(__file__))) # to print file name
    #csv2json()
    #readcsvjson()
    #csvdictreader()
    #xls2json()
    
    #getvaluefor = read_input('ogg_userdata_test.json')   
    '''
    getvaluefor = read_input('ogg_userdata.json')   
    src_ggadmin_user_name_v = getvaluefor['src_ggadmin_user_name']
    src_ggadmin_user_pwd_v = getvaluefor['src_ggadmin_user_pwd']
    src_schema_toreplicate_v = getvaluefor['src_schema_toreplicate']
    src_tns_name_v = getvaluefor['src_tns_name']
    print(src_ggadmin_user_name_v,'-',src_ggadmin_user_pwd_v)
    src_ddl_user_name_v = getvaluefor['src_ddl_user_name']
    src_ddl_user_pwd_v  = getvaluefor['src_ddl_user_pwd']
    src_ddl_user_tbs_name_v = getvaluefor['src_ddl_user_tbs_name']
    src_ddl_user_tbs_df_v = getvaluefor['src_ddl_user_tbs_df']
    src_schema_toreplicate_pwd_v = getvaluefor['src_schema_toreplicate_pwd']
    src_table_toreplicate_v = getvaluefor['src_table_toreplicate']
    src_tns_name_v = getvaluefor['src_tns_name']
    src_host_port_v = getvaluefor['src_host_port']
    src_host_ip_v = getvaluefor['src_host_ip']
    src_service_name_v = getvaluefor['src_service_name']
    tgt_ggadmin_user_name_v = getvaluefor['tgt_ggadmin_user_name']
    tgt_tns_name_v = getvaluefor['tgt_tns_name']
    tgt_ggadmin_user_pwd_v = getvaluefor['tgt_ggadmin_user_pwd']
    tgt_schema_tomap_v =  getvaluefor['tgt_schema_tomap']
    readjsonwitharg('ogg_config_test.json')
    #readjsonwitharg('ogg_config.json')
    '''
    
    #txt2xls()

    conn = srcdbconnect()
    vcur = conn.cursor()   
    fn_srcdb_check_dblogmode(vcur)    